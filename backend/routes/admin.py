from flask import Blueprint, jsonify, request, send_file, session
from models import db, Ruta, Etapa, Usuario, RutaFija
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.security import generate_password_hash
from datetime import datetime, timedelta
import io
import os

admin_bp = Blueprint('admin', __name__)


@admin_bp.before_request
def requerir_admin():
    if not session.get('usuario_id'):
        return jsonify({'error': 'No autorizado'}), 401
    if session.get('rol') not in ('admin', 'supervisor'):
        return jsonify({'error': 'No autorizado'}), 403


def _requiere_admin():
    if session.get('rol') != 'admin':
        return jsonify({'error': 'Solo administradores pueden ejecutar esta accion'}), 403
    return None


def _fmt_segundos(segundos):
    segundos = int(segundos or 0)
    return f"{segundos // 3600:02d}:{(segundos % 3600) // 60:02d}:{segundos % 60:02d}"


def _percentil(valores, percentil):
    if not valores:
        return 0
    ordenados = sorted(valores)
    posicion = (len(ordenados) - 1) * percentil
    inferior = int(posicion)
    superior = min(inferior + 1, len(ordenados) - 1)
    peso_superior = posicion - inferior
    return int(ordenados[inferior] * (1 - peso_superior) + ordenados[superior] * peso_superior)


def _ruta_fija(numero_ruta):
    return RutaFija.query.filter_by(numero_ruta=numero_ruta).first()


def _ruta_payload(ruta):
    ruta_fija = _ruta_fija(ruta.numero_ruta)
    etapas_info = []
    tiempo_total = 0

    for etapa in sorted(ruta.etapas, key=lambda x: x.orden):
        if etapa.nombre == 'Impresión de Previsita':
            continue
        tiempo = etapa.duracion_segundos or 0
        tiempo_total += tiempo
        etapas_info.append({
            'nombre': etapa.nombre,
            'duracion_segundos': tiempo,
            'duracion_formateada': etapa.get_duracion_formateada(),
            'completada': etapa.completada
        })

    return {
        'id': ruta.id,
        'numero_ruta': ruta.numero_ruta,
        'usuario': ruta.usuario.nombre if ruta.usuario else 'Sin asignar',
        'fecha': ruta.fecha.isoformat(),
        'estado': ruta.estado,
        'etapas': etapas_info,
        'tiempo_total_segundos': tiempo_total,
        'tiempo_total_formateado': _fmt_segundos(tiempo_total),
        'notas': ruta.notas,
        'supervisor': ruta_fija.supervisor if ruta_fija else 'Sin supervisor',
        'contratista': ruta_fija.contratista if ruta_fija else 'Sin contratista',
        'canal': ruta_fija.canal if ruta_fija else '-',
        'estatus_fijo': ruta_fija.estatus if ruta_fija else '-'
    }


def _duracion_etapa(ruta, nombre):
    etapa = next((etapa for etapa in ruta.etapas if etapa.nombre == nombre), None)
    return etapa.duracion_segundos if etapa and etapa.duracion_segundos is not None else None


def _aplicar_filtros_rutas(query):
    estado = request.args.get('estado')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_final = request.args.get('fecha_final')
    supervisor = request.args.get('supervisor')
    contratista = request.args.get('contratista')

    if estado:
        query = query.filter(Ruta.estado == estado)

    if fecha_inicio:
        fecha = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        query = query.filter(Ruta.fecha >= fecha)

    if fecha_final:
        fecha = datetime.strptime(fecha_final, '%Y-%m-%d')
        fecha_siguiente = fecha + timedelta(days=1)
        query = query.filter(Ruta.fecha < fecha_siguiente)

    if supervisor:
        rutas_supervisor = db.session.query(RutaFija.numero_ruta).filter(RutaFija.supervisor == supervisor)
        query = query.filter(Ruta.numero_ruta.in_(rutas_supervisor))

    if contratista:
        rutas_contratista = db.session.query(RutaFija.numero_ruta).filter(RutaFija.contratista == contratista)
        query = query.filter(Ruta.numero_ruta.in_(rutas_contratista))

    return query


@admin_bp.route('/rutas-con-etapas', methods=['GET'])
def obtener_rutas_con_etapas():
    """Obtiene todas las rutas con sus etapas para admin"""
    filtro_usuario = request.args.get('usuario_id', type=int)
    filtro_estado = request.args.get('estado')
    filtro_fecha = request.args.get('fecha')
    filtro_supervisor = request.args.get('supervisor')
    filtro_contratista = request.args.get('contratista')
    
    query = Ruta.query
    
    if filtro_usuario:
        query = query.filter_by(usuario_id=filtro_usuario)
    if filtro_estado:
        query = query.filter_by(estado=filtro_estado)
    if filtro_fecha:
        fecha_obj = datetime.strptime(filtro_fecha, '%Y-%m-%d')
        fecha_siguiente = fecha_obj + timedelta(days=1)
        query = query.filter(Ruta.fecha >= fecha_obj, Ruta.fecha < fecha_siguiente)
    if filtro_supervisor:
        rutas_supervisor = db.session.query(RutaFija.numero_ruta).filter(RutaFija.supervisor == filtro_supervisor)
        query = query.filter(Ruta.numero_ruta.in_(rutas_supervisor))
    if filtro_contratista:
        rutas_contratista = db.session.query(RutaFija.numero_ruta).filter(RutaFija.contratista == filtro_contratista)
        query = query.filter(Ruta.numero_ruta.in_(rutas_contratista))
    
    rutas = query.order_by(Ruta.fecha.desc()).all()
    
    return jsonify([_ruta_payload(ruta) for ruta in rutas])


@admin_bp.route('/descargar-excel', methods=['GET'])
def descargar_excel():
    """Descarga las rutas en formato Excel"""
    filtro_usuario = request.args.get('usuario_id', type=int)
    filtro_estado = request.args.get('estado')
    filtro_fecha = request.args.get('fecha')
    filtro_supervisor = request.args.get('supervisor')
    filtro_contratista = request.args.get('contratista')
    
    query = Ruta.query
    
    if filtro_usuario:
        query = query.filter_by(usuario_id=filtro_usuario)
    if filtro_estado:
        query = query.filter_by(estado=filtro_estado)
    if filtro_fecha:
        fecha_obj = datetime.strptime(filtro_fecha, '%Y-%m-%d')
        fecha_siguiente = fecha_obj + timedelta(days=1)
        query = query.filter(Ruta.fecha >= fecha_obj, Ruta.fecha < fecha_siguiente)
    if filtro_supervisor:
        rutas_supervisor = db.session.query(RutaFija.numero_ruta).filter(RutaFija.supervisor == filtro_supervisor)
        query = query.filter(Ruta.numero_ruta.in_(rutas_supervisor))
    if filtro_contratista:
        rutas_contratista = db.session.query(RutaFija.numero_ruta).filter(RutaFija.contratista == filtro_contratista)
        query = query.filter(Ruta.numero_ruta.in_(rutas_contratista))
    
    rutas = query.order_by(Ruta.fecha.desc()).all()
    
    # Crear libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Rutas"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['ID', 'Ruta', 'Supervisor', 'Contratista', 'Canal', 'Usuario', 'Fecha', 'Estado', 'Matinal', 'Tiempo en Fila', 'Conteo', 'Check Salida', 'Pánico', 'Tiempo Total', 'Notas']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    row = 2
    for ruta in rutas:
        ruta_fija = _ruta_fija(ruta.numero_ruta)
        # Obtener duraciones de etapas
        etapas_duracion = {}
        for etapa in ruta.etapas:
            etapas_duracion[etapa.nombre] = etapa.get_duracion_formateada() if etapa.duracion_segundos else "-"
        
        # Calcular tiempo total
        tiempo_total = sum(etapa.duracion_segundos or 0 for etapa in ruta.etapas)
        tiempo_total_fmt = f"{tiempo_total // 3600:02d}:{(tiempo_total % 3600) // 60:02d}:{tiempo_total % 60:02d}"
        
        datos = [
            ruta.id,
            ruta.numero_ruta,
            ruta_fija.supervisor if ruta_fija else '',
            ruta_fija.contratista if ruta_fija else '',
            ruta_fija.canal if ruta_fija else '',
            ruta.usuario.nombre if ruta.usuario else 'Sin asignar',
            ruta.fecha.strftime('%Y-%m-%d %H:%M:%S'),
            ruta.estado,
            etapas_duracion.get('Matinal', '-'),
            etapas_duracion.get('Conteo de Carga', '-'),
            etapas_duracion.get('Check de Salida', '-'),
            etapas_duracion.get('Botón de Pánico', '-'),
            etapas_duracion.get('Tiempo en Fila', '-'),
            tiempo_total_fmt,
            ruta.notas or ''
        ]
        
        for col, valor in enumerate(datos, 1):
            cell = ws.cell(row=row, column=col, value=valor)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 12
    ws.column_dimensions['K'].width = 15
    ws.column_dimensions['L'].width = 15
    ws.column_dimensions['M'].width = 12
    ws.column_dimensions['N'].width = 15
    ws.column_dimensions['O'].width = 20
    
    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'rutas_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )


@admin_bp.route('/rutas/<int:ruta_id>', methods=['PUT'])
def editar_ruta_admin(ruta_id):
    """Edita cabecera y etapas de una ruta. Accesible para admin y supervisor."""
    ruta = Ruta.query.get_or_404(ruta_id)
    data = request.get_json(silent=True) or {}

    # ── Campos de cabecera ──────────────────────────────────────────
    if 'numero_ruta' in data:
        ruta.numero_ruta = str(data['numero_ruta']).strip().upper()

    if 'estado' in data:
        if data['estado'] not in ('activa', 'completada', 'cancelada'):
            return jsonify({'error': 'Estado inválido. Use: activa, completada, cancelada'}), 400
        ruta.estado = data['estado']

    if 'notas' in data:
        ruta.notas = (data['notas'] or '').strip() or None

    if 'fecha' in data:
        for fmt in ('%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d'):
            try:
                ruta.fecha = datetime.strptime(data['fecha'], fmt)
                break
            except ValueError:
                continue
        else:
            return jsonify({'error': 'Formato de fecha inválido. Use YYYY-MM-DD o YYYY-MM-DDTHH:MM:SS'}), 400

    if 'usuario_id' in data:
        nuevo_usuario = Usuario.query.get(data['usuario_id'])
        if not nuevo_usuario:
            return jsonify({'error': 'Usuario no encontrado'}), 404
        ruta.usuario_id = nuevo_usuario.id

    # ── Etapas ──────────────────────────────────────────────────────
    # Lista parcial: [{ "nombre": "Matinal", "duracion_segundos": 300, "completada": true }, ...]
    # Solo se actualizan las etapas incluidas; las demás no se tocan.
    etapas_payload = data.get('etapas')
    if etapas_payload is not None:
        if not isinstance(etapas_payload, list):
            return jsonify({'error': 'El campo etapas debe ser una lista'}), 400

        etapas_por_nombre = {etapa.nombre: etapa for etapa in ruta.etapas}

        for item in etapas_payload:
            nombre = (item.get('nombre') or '').strip()
            if not nombre:
                continue

            etapa = etapas_por_nombre.get(nombre)
            if etapa is None:
                orden_max = max((e.orden for e in ruta.etapas), default=0)
                etapa = Etapa(ruta_id=ruta.id, nombre=nombre, orden=orden_max + 1)
                db.session.add(etapa)
                etapas_por_nombre[nombre] = etapa

            if 'duracion_segundos' in item:
                valor = item['duracion_segundos']
                if valor is not None and not isinstance(valor, (int, float)):
                    return jsonify({'error': f'duracion_segundos inválido en etapa "{nombre}"'}), 400
                etapa.duracion_segundos = int(valor) if valor is not None else None

            if 'completada' in item:
                etapa.completada = bool(item['completada'])

    db.session.commit()
    return jsonify({'mensaje': 'Ruta actualizada correctamente', 'ruta': _ruta_payload(ruta)})


@admin_bp.route('/rutas/<int:ruta_id>', methods=['DELETE'])
def eliminar_ruta_admin(ruta_id):
    """Elimina una ruta medida con sus etapas"""
    autorizado = _requiere_admin()
    if autorizado:
        return autorizado

    ruta = Ruta.query.get_or_404(ruta_id)
    db.session.delete(ruta)
    db.session.commit()
    return jsonify({'mensaje': 'Ruta eliminada'})


@admin_bp.route('/limpiar-datos', methods=['POST'])
def limpiar_datos():
    """Elimina datos transaccionales de prueba sin borrar usuarios ni rutas fijas"""
    autorizado = _requiere_admin()
    if autorizado:
        return autorizado

    data = request.get_json(silent=True) or {}
    confirmacion = data.get('confirmacion', '')
    if confirmacion != 'BORRAR':
        return jsonify({'error': 'Confirmacion requerida'}), 400

    rutas = Ruta.query.all()
    total = len(rutas)
    for ruta in rutas:
        db.session.delete(ruta)
    db.session.commit()
    return jsonify({'mensaje': 'Datos de rutas eliminados', 'rutas_eliminadas': total})


@admin_bp.route('/importar-rutas-fijas', methods=['GET', 'POST'])
def importar_rutas_fijas():
    """Importa la base de supervisores/contratistas desde Rutas fijas Mayo.xlsx"""
    autorizado = _requiere_admin()
    if autorizado:
        return autorizado

    base_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..'))
    archivo = os.path.join(base_dir, 'Rutas fijas Mayo.xlsx')
    if not os.path.exists(archivo):
        return jsonify({'error': 'No se encontro Rutas fijas Mayo.xlsx'}), 404

    wb = load_workbook(archivo, data_only=True)
    ws = wb['Hoja3'] if 'Hoja3' in wb.sheetnames else wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor:
            headers[str(valor).strip().upper()] = col

    requeridas = ['RUTA', 'CONTRATISTA', 'ESTATUS', 'CANAL', 'SUPERVISOR']
    faltantes = [col for col in requeridas if col not in headers]
    if faltantes:
        return jsonify({'error': f'Columnas faltantes: {", ".join(faltantes)}'}), 400

    importadas = 0
    actualizadas = 0
    for row in range(2, ws.max_row + 1):
        numero = ws.cell(row=row, column=headers['RUTA']).value
        if not numero:
            continue

        numero_ruta = str(numero).strip().upper()
        fija = RutaFija.query.filter_by(numero_ruta=numero_ruta).first()
        if not fija:
            fija = RutaFija(numero_ruta=numero_ruta)
            db.session.add(fija)
            importadas += 1
        else:
            actualizadas += 1

        fija.contratista = str(ws.cell(row=row, column=headers['CONTRATISTA']).value or '').strip()
        fija.estatus = str(ws.cell(row=row, column=headers['ESTATUS']).value or '').strip()
        fija.canal = str(ws.cell(row=row, column=headers['CANAL']).value or '').strip()
        fija.supervisor = str(ws.cell(row=row, column=headers['SUPERVISOR']).value or '').strip()
        fija.origen = os.path.basename(archivo)
        fija.fecha_importacion = datetime.utcnow()

    db.session.commit()
    return jsonify({
        'mensaje': 'Rutas fijas importadas',
        'importadas': importadas,
        'actualizadas': actualizadas,
        'total': RutaFija.query.count()
    })


@admin_bp.route('/rutas-fijas', methods=['GET'])
def obtener_rutas_fijas():
    rutas = RutaFija.query.order_by(RutaFija.supervisor, RutaFija.numero_ruta).all()
    return jsonify([{
        'id': ruta.id,
        'numero_ruta': ruta.numero_ruta,
        'contratista': ruta.contratista,
        'estatus': ruta.estatus,
        'canal': ruta.canal,
        'supervisor': ruta.supervisor,
        'origen': ruta.origen
    } for ruta in rutas])


@admin_bp.route('/opciones-filtros', methods=['GET'])
def opciones_filtros():
    supervisores = [item[0] for item in db.session.query(RutaFija.supervisor).filter(RutaFija.supervisor != '').distinct().order_by(RutaFija.supervisor).all()]
    contratistas = [item[0] for item in db.session.query(RutaFija.contratista).filter(RutaFija.contratista != '').distinct().order_by(RutaFija.contratista).all()]
    return jsonify({'supervisores': supervisores, 'contratistas': contratistas})


@admin_bp.route('/usuarios', methods=['GET'])
def listar_usuarios():
    autorizado = _requiere_admin()
    if autorizado:
        return autorizado

    usuarios = Usuario.query.order_by(Usuario.nombre).all()
    return jsonify([{
        'id': usuario.id,
        'nombre': usuario.nombre,
        'email': usuario.email,
        'rol': usuario.rol or ('admin' if usuario.es_admin else 'operador'),
        'activo': usuario.activo,
        'fecha_creacion': usuario.fecha_creacion.isoformat() if usuario.fecha_creacion else None
    } for usuario in usuarios])


@admin_bp.route('/usuarios', methods=['POST'])
def crear_usuario():
    autorizado = _requiere_admin()
    if autorizado:
        return autorizado

    data = request.get_json(silent=True) or {}
    nombre = (data.get('nombre') or '').strip()
    email = (data.get('email') or '').strip().lower()
    password = data.get('password') or ''
    rol = data.get('rol') or 'operador'

    if rol not in ('admin', 'supervisor', 'operador'):
        return jsonify({'error': 'Rol invalido'}), 400
    if not nombre or not email or not password:
        return jsonify({'error': 'Nombre, email y contrasena son obligatorios'}), 400
    if Usuario.query.filter_by(email=email).first():
        return jsonify({'error': 'Ya existe un usuario con ese email'}), 400

    usuario = Usuario(
        nombre=nombre,
        email=email,
        password=generate_password_hash(password),
        rol=rol,
        es_admin=rol == 'admin',
        activo=bool(data.get('activo', True))
    )
    db.session.add(usuario)
    db.session.commit()
    return jsonify({'mensaje': 'Usuario creado', 'id': usuario.id}), 201


@admin_bp.route('/usuarios/<int:usuario_id>', methods=['PUT'])
def actualizar_usuario(usuario_id):
    autorizado = _requiere_admin()
    if autorizado:
        return autorizado

    usuario = Usuario.query.get_or_404(usuario_id)
    data = request.get_json(silent=True) or {}

    if 'nombre' in data:
        usuario.nombre = (data.get('nombre') or '').strip()
    if 'email' in data:
        email = (data.get('email') or '').strip().lower()
        existente = Usuario.query.filter(Usuario.email == email, Usuario.id != usuario.id).first()
        if existente:
            return jsonify({'error': 'Ya existe un usuario con ese email'}), 400
        usuario.email = email
    if 'rol' in data:
        rol = data.get('rol')
        if rol not in ('admin', 'supervisor', 'operador'):
            return jsonify({'error': 'Rol invalido'}), 400
        usuario.rol = rol
        usuario.es_admin = rol == 'admin'
    if 'activo' in data:
        usuario.activo = bool(data.get('activo'))
    if data.get('password'):
        usuario.password = generate_password_hash(data['password'])

    db.session.commit()
    return jsonify({'mensaje': 'Usuario actualizado'})


@admin_bp.route('/usuarios/<int:usuario_id>', methods=['DELETE'])
def eliminar_usuario(usuario_id):
    autorizado = _requiere_admin()
    if autorizado:
        return autorizado

    if usuario_id == session.get('usuario_id'):
        return jsonify({'error': 'No puedes eliminar tu propio usuario'}), 400

    usuario = Usuario.query.get_or_404(usuario_id)
    if usuario.rutas:
        return jsonify({'error': 'Este usuario tiene rutas asociadas. Desactivalo para conservar el historial.'}), 400

    db.session.delete(usuario)
    db.session.commit()
    return jsonify({'mensaje': 'Usuario eliminado'})


@admin_bp.route('/estadisticas-general', methods=['GET'])
def estadisticas_general():
    """Estadísticas generales del sistema"""
    dias = request.args.get('dias', default=30, type=int)
    fecha_inicio_str = request.args.get('fecha_inicio')
    fecha_final_str = request.args.get('fecha_final')
    objetivo_total = request.args.get('objetivo_total_min', default=45, type=int) * 60
    objetivo_fila = request.args.get('objetivo_fila_min', default=5, type=int) * 60

    fecha_inicio = None
    if fecha_inicio_str:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
    elif dias:
        fecha_inicio = datetime.utcnow() - timedelta(days=dias)

    query = Ruta.query
    query = _aplicar_filtros_rutas(query)

    if fecha_inicio:
        query = query.filter(Ruta.fecha >= fecha_inicio)
    if fecha_final_str:
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Ruta.fecha < fecha_final)

    rutas_periodo = query.all()
    rutas_totales = len(rutas_periodo)
    rutas_completadas = sum(1 for ruta in rutas_periodo if ruta.estado == 'completada')
    rutas_activas = sum(1 for ruta in rutas_periodo if ruta.estado == 'activa')
    rutas_canceladas = sum(1 for ruta in rutas_periodo if ruta.estado == 'cancelada')

    usuarios_totales = Usuario.query.filter(Usuario.activo == True).count()

    etapas_stats = {}
    for ruta in rutas_periodo:
        for etapa in ruta.etapas:
            if etapa.duracion_segundos is None:
                continue
            nombre = etapa.nombre
            if nombre not in etapas_stats:
                etapas_stats[nombre] = []
            etapas_stats[nombre].append(etapa.duracion_segundos)

    etapas_promedio = {}
    for nombre, tiempos in etapas_stats.items():
        if nombre == 'Impresión de Previsita':
            continue
        promedio = sum(tiempos) / len(tiempos)
        etapas_promedio[nombre] = {
            'promedio': int(promedio),
            'promedio_fmt': f"{int(promedio) // 60:02d}:{int(promedio) % 60:02d}",
            'minimo': min(tiempos),
            'maximo': max(tiempos),
            'cantidad': len(tiempos)
        }

    tiempos_totales = []
    for ruta in rutas_periodo:
        if ruta.estado == 'completada':
            total = sum(etapa.duracion_segundos or 0 for etapa in ruta.etapas)
            if total:
                tiempos_totales.append((ruta, total))

    tiempos_fila = []
    for ruta in rutas_periodo:
        fila = _duracion_etapa(ruta, 'Tiempo en Fila')
        if fila is not None:
            tiempos_fila.append(fila)

    valores_totales = [total for _, total in tiempos_totales]
    promedio_total = int(sum(valores_totales) / len(valores_totales)) if valores_totales else 0
    ruta_mas_rapida = min(tiempos_totales, key=lambda item: item[1], default=None)
    ruta_mas_lenta = max(tiempos_totales, key=lambda item: item[1], default=None)
    cuello_botella = max(etapas_promedio.items(), key=lambda item: item[1]['promedio'], default=None)

    promedio_fila = int(sum(tiempos_fila) / len(tiempos_fila)) if tiempos_fila else 0
    tiempo_fila_maximo = max(tiempos_fila) if tiempos_fila else 0
    rutas_con_tiempo_fila = len(tiempos_fila)

    estados = {
        'activa': rutas_activas,
        'completada': rutas_completadas,
        'cancelada': rutas_canceladas
    }
    p50_total = _percentil(valores_totales, 0.50)
    p90_total = _percentil(valores_totales, 0.90)
    p95_total = _percentil(valores_totales, 0.95)
    p50_fila = _percentil(tiempos_fila, 0.50)
    p90_fila = _percentil(tiempos_fila, 0.90)
    p95_fila = _percentil(tiempos_fila, 0.95)
    rutas_en_objetivo_total = sum(1 for total in valores_totales if total <= objetivo_total)
    rutas_en_objetivo_fila = sum(1 for fila in tiempos_fila if fila <= objetivo_fila)
    cobertura_fila = (rutas_con_tiempo_fila / rutas_totales * 100) if rutas_totales else 0
    cumplimiento_total = (rutas_en_objetivo_total / len(valores_totales) * 100) if valores_totales else 0
    cumplimiento_fila = (rutas_en_objetivo_fila / len(tiempos_fila) * 100) if tiempos_fila else 0

    return jsonify({
        'rutas_totales': rutas_totales,
        'rutas_completadas': rutas_completadas,
        'rutas_activas': rutas_activas,
        'rutas_canceladas': rutas_canceladas,
        'tasa_completacion': f"{(rutas_completadas / rutas_totales * 100) if rutas_totales > 0 else 0:.1f}%",
        'usuarios_activos': usuarios_totales,
        'tiempo_promedio_total': promedio_total,
        'tiempo_promedio_total_fmt': f"{promedio_total // 3600:02d}:{(promedio_total % 3600) // 60:02d}:{promedio_total % 60:02d}",
        'ruta_mas_rapida': {
            'numero_ruta': ruta_mas_rapida[0].numero_ruta,
            'tiempo_segundos': ruta_mas_rapida[1],
            'tiempo_fmt': f"{ruta_mas_rapida[1] // 3600:02d}:{(ruta_mas_rapida[1] % 3600) // 60:02d}:{ruta_mas_rapida[1] % 60:02d}"
        } if ruta_mas_rapida else None,
        'ruta_mas_lenta': {
            'numero_ruta': ruta_mas_lenta[0].numero_ruta,
            'tiempo_segundos': ruta_mas_lenta[1],
            'tiempo_fmt': f"{ruta_mas_lenta[1] // 3600:02d}:{(ruta_mas_lenta[1] % 3600) // 60:02d}:{ruta_mas_lenta[1] % 60:02d}"
        } if ruta_mas_lenta else None,
        'cuello_botella': {
            'nombre': cuello_botella[0],
            'promedio': cuello_botella[1]['promedio'],
            'promedio_fmt': cuello_botella[1]['promedio_fmt']
        } if cuello_botella else None,
        'estados': estados,
        'etapas_promedio': etapas_promedio,
        'tiempo_en_fila_promedio': promedio_fila,
        'tiempo_en_fila_promedio_fmt': f"{promedio_fila // 3600:02d}:{(promedio_fila % 3600) // 60:02d}:{promedio_fila % 60:02d}",
        'tiempo_en_fila_maximo': tiempo_fila_maximo,
        'rutas_con_tiempo_en_fila': rutas_con_tiempo_fila,
        'rutas_sin_tiempo_en_fila': rutas_totales - rutas_con_tiempo_fila,
        'cobertura_tiempo_fila_pct': round(cobertura_fila, 1),
        'p50_total': p50_total,
        'p50_total_fmt': _fmt_segundos(p50_total),
        'p90_total': p90_total,
        'p90_total_fmt': _fmt_segundos(p90_total),
        'p95_total': p95_total,
        'p95_total_fmt': _fmt_segundos(p95_total),
        'p50_fila': p50_fila,
        'p50_fila_fmt': _fmt_segundos(p50_fila),
        'p90_fila': p90_fila,
        'p90_fila_fmt': _fmt_segundos(p90_fila),
        'p95_fila': p95_fila,
        'p95_fila_fmt': _fmt_segundos(p95_fila),
        'objetivo_total_segundos': objetivo_total,
        'objetivo_total_fmt': _fmt_segundos(objetivo_total),
        'rutas_en_objetivo_total': rutas_en_objetivo_total,
        'cumplimiento_objetivo_total_pct': round(cumplimiento_total, 1),
        'objetivo_fila_segundos': objetivo_fila,
        'objetivo_fila_fmt': _fmt_segundos(objetivo_fila),
        'rutas_en_objetivo_fila': rutas_en_objetivo_fila,
        'cumplimiento_objetivo_fila_pct': round(cumplimiento_fila, 1),
        'periodo_dias': dias
    })


@admin_bp.route('/analisis-avanzado', methods=['GET'])
def analisis_avanzado():
    """Analisis operativo para fortalecer el dashboard"""
    dias = request.args.get('dias', default=30, type=int)
    fecha_inicio_str = request.args.get('fecha_inicio')
    fecha_final_str = request.args.get('fecha_final')
    objetivo_total = request.args.get('objetivo_total_min', default=45, type=int) * 60
    objetivo_fila = request.args.get('objetivo_fila_min', default=5, type=int) * 60

    fecha_inicio = None
    if fecha_inicio_str:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d')
    elif dias:
        fecha_inicio = datetime.utcnow() - timedelta(days=dias)

    query = Ruta.query
    query = _aplicar_filtros_rutas(query)

    if fecha_inicio:
        query = query.filter(Ruta.fecha >= fecha_inicio)
    if fecha_final_str:
        fecha_final = datetime.strptime(fecha_final_str, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Ruta.fecha < fecha_final)

    rutas = query.order_by(Ruta.fecha.desc()).all()
    por_hora = {str(hora).zfill(2): 0 for hora in range(24)}
    ranking_etapas = {}
    supervisores = {}
    contratistas = {}
    canales = {}
    tendencia_diaria = {}
    distribucion_total = {
        '0-10 min': 0,
        '10-20 min': 0,
        '20-30 min': 0,
        '30-45 min': 0,
        '45+ min': 0
    }
    rutas_problematicas = []
    rutas_lentas = []
    rutas_cola = []

    for ruta in rutas:
        por_hora[str(ruta.fecha.hour).zfill(2)] += 1
        total = sum(etapa.duracion_segundos or 0 for etapa in ruta.etapas)
        pendientes = sum(1 for etapa in ruta.etapas if not etapa.completada)
        tiempo_fila = _duracion_etapa(ruta, 'Tiempo en Fila')
        ruta_fija = _ruta_fija(ruta.numero_ruta)
        supervisor = ruta_fija.supervisor if ruta_fija and ruta_fija.supervisor else 'Sin supervisor'
        contratista = ruta_fija.contratista if ruta_fija and ruta_fija.contratista else 'Sin contratista'
        canal = ruta_fija.canal if ruta_fija and ruta_fija.canal else 'Sin canal'
        fecha_key = ruta.fecha.strftime('%Y-%m-%d')

        if fecha_key not in tendencia_diaria:
            tendencia_diaria[fecha_key] = {'total': 0, 'completadas': 0, 'canceladas': 0, 'en_objetivo': 0}
        tendencia_diaria[fecha_key]['total'] += 1
        if ruta.estado == 'completada':
            tendencia_diaria[fecha_key]['completadas'] += 1
        if ruta.estado == 'cancelada':
            tendencia_diaria[fecha_key]['canceladas'] += 1
        if ruta.estado == 'completada' and total and total <= objetivo_total:
            tendencia_diaria[fecha_key]['en_objetivo'] += 1

        if canal not in canales:
            canales[canal] = {'rutas': 0, 'completadas': 0}
        canales[canal]['rutas'] += 1
        if ruta.estado == 'completada':
            canales[canal]['completadas'] += 1

        if ruta.estado == 'completada' and total:
            minutos = total / 60
            if minutos < 10:
                distribucion_total['0-10 min'] += 1
            elif minutos < 20:
                distribucion_total['10-20 min'] += 1
            elif minutos < 30:
                distribucion_total['20-30 min'] += 1
            elif minutos < 45:
                distribucion_total['30-45 min'] += 1
            else:
                distribucion_total['45+ min'] += 1
            rutas_lentas.append({
                'numero_ruta': ruta.numero_ruta,
                'estado': ruta.estado,
                'tiempo_segundos': total,
                'tiempo_fmt': _fmt_segundos(total),
                'supervisor': supervisor,
                'contratista': contratista
            })

        if tiempo_fila is not None:
            rutas_cola.append({
                'numero_ruta': ruta.numero_ruta,
                'estado': ruta.estado,
                'tiempo_segundos': tiempo_fila,
                'tiempo_fmt': _fmt_segundos(tiempo_fila),
                'supervisor': supervisor
            })

        if ruta.estado != 'completada' and pendientes > 0:
            rutas_problematicas.append({
                'id': ruta.id,
                'numero_ruta': ruta.numero_ruta,
                'estado': ruta.estado,
                'pendientes': pendientes,
                'supervisor': supervisor,
                'contratista': contratista,
                'motivo': 'Etapas pendientes',
                'tiempo_fmt': _fmt_segundos(total)
            })
        elif ruta.estado == 'completada' and total and total > objetivo_total:
            rutas_problematicas.append({
                'id': ruta.id,
                'numero_ruta': ruta.numero_ruta,
                'estado': ruta.estado,
                'pendientes': pendientes,
                'supervisor': supervisor,
                'contratista': contratista,
                'motivo': 'Fuera de objetivo',
                'tiempo_fmt': _fmt_segundos(total)
            })
        elif tiempo_fila is not None and tiempo_fila > objetivo_fila:
            rutas_problematicas.append({
                'id': ruta.id,
                'numero_ruta': ruta.numero_ruta,
                'estado': ruta.estado,
                'pendientes': pendientes,
                'supervisor': supervisor,
                'contratista': contratista,
                'motivo': 'Tiempo en fila alto',
                'tiempo_fmt': _fmt_segundos(tiempo_fila)
            })
        elif ruta.estado == 'cancelada':
            rutas_problematicas.append({
                'id': ruta.id,
                'numero_ruta': ruta.numero_ruta,
                'estado': ruta.estado,
                'pendientes': pendientes,
                'supervisor': supervisor,
                'contratista': contratista,
                'motivo': 'Ruta cancelada',
                'tiempo_fmt': _fmt_segundos(total)
            })

        for bucket, nombre in ((supervisores, supervisor), (contratistas, contratista)):
            if nombre not in bucket:
                bucket[nombre] = {
                    'rutas': 0,
                    'completadas': 0,
                    'tiempo_total': 0,
                    'con_tiempo': 0,
                    'cola_total': 0,
                    'cola_count': 0,
                    'tiempos_totales': [],
                    'tiempos_cola': [],
                    'etapas': {}
                }
            bucket[nombre]['rutas'] += 1
            if ruta.estado == 'completada':
                bucket[nombre]['completadas'] += 1
            if total:
                bucket[nombre]['tiempo_total'] += total
                bucket[nombre]['con_tiempo'] += 1
                bucket[nombre]['tiempos_totales'].append(total)
            if tiempo_fila is not None:
                bucket[nombre]['cola_total'] += tiempo_fila
                bucket[nombre]['cola_count'] += 1
                bucket[nombre]['tiempos_cola'].append(tiempo_fila)

            for etapa in ruta.etapas:
                if etapa.duracion_segundos is None:
                    continue
                etapa_nombre = etapa.nombre
                if etapa_nombre not in bucket[nombre]['etapas']:
                    bucket[nombre]['etapas'][etapa_nombre] = {'total': 0, 'count': 0}
                bucket[nombre]['etapas'][etapa_nombre]['total'] += etapa.duracion_segundos
                bucket[nombre]['etapas'][etapa_nombre]['count'] += 1

        for etapa in ruta.etapas:
            if etapa.duracion_segundos is None:
                continue
            if etapa.nombre not in ranking_etapas:
                ranking_etapas[etapa.nombre] = {'total': 0, 'cantidad': 0, 'maximo': 0}
            ranking_etapas[etapa.nombre]['total'] += etapa.duracion_segundos
            ranking_etapas[etapa.nombre]['cantidad'] += 1
            ranking_etapas[etapa.nombre]['maximo'] = max(ranking_etapas[etapa.nombre]['maximo'], etapa.duracion_segundos)

    etapas_ordenadas = []
    for nombre, data in ranking_etapas.items():
        promedio = int(data['total'] / data['cantidad']) if data['cantidad'] else 0
        etapas_ordenadas.append({
            'nombre': nombre,
            'promedio': promedio,
            'promedio_fmt': f"{promedio // 60:02d}:{promedio % 60:02d}",
            'maximo': data['maximo'],
            'maximo_fmt': f"{data['maximo'] // 60:02d}:{data['maximo'] % 60:02d}",
            'cantidad': data['cantidad']
        })

    etapas_ordenadas.sort(key=lambda item: item['promedio'], reverse=True)

    def resumir_bucket(bucket):
        resumen = []
        for nombre, data in bucket.items():
            promedio = int(data['tiempo_total'] / data['con_tiempo']) if data['con_tiempo'] else 0
            promedio_cola = int(data['cola_total'] / data['cola_count']) if data['cola_count'] else 0
            etapas = []
            for etapa_nombre, etapa_data in data['etapas'].items():
                promedio_etapa = int(etapa_data['total'] / etapa_data['count']) if etapa_data['count'] else 0
                etapas.append({
                    'nombre': etapa_nombre,
                    'promedio': promedio_etapa,
                    'promedio_fmt': _fmt_segundos(promedio_etapa),
                    'cantidad': etapa_data['count']
                })
            etapas.sort(key=lambda item: item['promedio'], reverse=True)
            p50_total = _percentil(data['tiempos_totales'], 0.50)
            p90_total = _percentil(data['tiempos_totales'], 0.90)
            p95_total = _percentil(data['tiempos_totales'], 0.95)
            p50_cola = _percentil(data['tiempos_cola'], 0.50)
            p90_cola = _percentil(data['tiempos_cola'], 0.90)
            p95_cola = _percentil(data['tiempos_cola'], 0.95)
            fuera_total = sum(1 for t in data['tiempos_totales'] if t > objetivo_total)
            fuera_cola = sum(1 for t in data['tiempos_cola'] if t > objetivo_fila)
            cumplimiento_total = round((1 - fuera_total / len(data['tiempos_totales']) * 100) if data['tiempos_totales'] else 0, 1)
            cumplimiento_cola = round((1 - fuera_cola / len(data['tiempos_cola']) * 100) if data['tiempos_cola'] else 0, 1)

            resumen.append({
                'nombre': nombre,
                'rutas': data['rutas'],
                'completadas': data['completadas'],
                'tasa': round((data['completadas'] / data['rutas'] * 100) if data['rutas'] else 0, 1),
                'promedio': promedio,
                'promedio_fmt': _fmt_segundos(promedio),
                'promedio_cola': promedio_cola,
                'promedio_cola_fmt': _fmt_segundos(promedio_cola),
                'cola_disponibles': data['cola_count'],
                'p50_total': p50_total,
                'p90_total': p90_total,
                'p95_total': p95_total,
                'p50_cola': p50_cola,
                'p90_cola': p90_cola,
                'p95_cola': p95_cola,
                'rutas_en_objetivo_total': len(data['tiempos_totales']) - fuera_total,
                'rutas_fuera_objetivo_total': fuera_total,
                'cumplimiento_objetivo_total_pct': cumplimiento_total,
                'rutas_en_objetivo_cola': len(data['tiempos_cola']) - fuera_cola,
                'rutas_fuera_objetivo_cola': fuera_cola,
                'cumplimiento_objetivo_cola_pct': cumplimiento_cola,
                'etapas': etapas
            })
        resumen.sort(key=lambda item: (item['rutas'], item['completadas']), reverse=True)
        return resumen

    def resumir_canales():
        resumen = []
        for nombre, data in canales.items():
            resumen.append({
                'nombre': nombre,
                'rutas': data['rutas'],
                'completadas': data['completadas'],
                'tasa': round((data['completadas'] / data['rutas'] * 100) if data['rutas'] else 0, 1)
            })
        resumen.sort(key=lambda item: item['rutas'], reverse=True)
        return resumen

    for data in tendencia_diaria.values():
        data['tasa_completacion'] = round((data['completadas'] / data['total'] * 100) if data['total'] else 0, 1)
        data['cumplimiento_objetivo'] = round((data['en_objetivo'] / data['completadas'] * 100) if data['completadas'] else 0, 1)

    rutas_lentas.sort(key=lambda item: item['tiempo_segundos'], reverse=True)
    rutas_cola.sort(key=lambda item: item['tiempo_segundos'], reverse=True)

    return jsonify({
        'por_hora': por_hora,
        'ranking_etapas': etapas_ordenadas,
        'supervisores': resumir_bucket(supervisores),
        'contratistas': resumir_bucket(contratistas),
        'canales': resumir_canales(),
        'tendencia_diaria': tendencia_diaria,
        'distribucion_total': distribucion_total,
        'top_rutas_lentas': rutas_lentas[:10],
        'top_rutas_cola': rutas_cola[:10],
        'rutas_problematicas': rutas_problematicas[:10]
    })
